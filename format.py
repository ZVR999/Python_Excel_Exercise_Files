# -*- coding: utf-8 -*-
"""
Created on Thu Jul 29 14:37:39 2021

@author: zachk
"""

from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

#create a workbook
wb = Workbook()
#set the active worksheet to this variable
ws = wb.active

#create values to work with
for i in range(1,20):
    ws.append(range(300))
    
#merge cells A1 to B5
ws.merge_cells('A1:B5')
#unmerge cells A1 to B5
ws.unmerge_cells('A1:B5')
#use flags to for selection
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

#This will place the merged cells starting with B2 in this variable
cell = ws['B2']
#Style the font of the 'cell' variable
cell.font = Font(color='00FF0000', size=20, italic=True)
#Change the value of the 'cell' variable aka B2
cell.value = 'Merged Cell'
#Change the alignment of the cell
cell.alignment = Alignment(horizontal='right', vertical='bottom')
#Apply a gradientfill to the merged cell
cell.fill = GradientFill(stop=('000000', 'FFFFFF'))
#Save the changes to a new xlsx file to see them
wb.save('text.xlsx')

#initialize the style
highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
bd = Side(style='thick', color='000000')
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill('solid', fgColor='FFFF00')

count = 0
#highlight cells starting at H1 and count to the right and down diagonally in sheet
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count  = count + 1
# wb.save('highlight.xlsx')

