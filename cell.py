# -*- coding: utf-8 -*-
"""
Created on Thu Jul 29 13:39:07 2021

@author: zachk
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('regions.xlsx')
ws = wb.active

cell_range = ws['A1':'C1']

# print(cell_range)

# set column 'C' to this variable
col_c = ws['C']
# display column 'C'
# print(col_c)

# set columns 'A' through 'C' to this variable
col_range = ws['A':'D']
# display the selected colonms
# print(col_range)

row_range = ws[1:5]
# print(row_range)

# iterate through rows 1 to 3 and display the location and value
#of every cell
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
            print(cell)
            print(cell.value)
            
            