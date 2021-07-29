# -*- coding: utf-8 -*-
"""
Created on Thu Jul 29 13:02:32 2021

@author: zachk
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()

ws = wb.active

#Create a new sheet called 'NewSheet' that will show up
#after the current sheet
ws1 = wb.create_sheet('NewSheet')

#The 0 is to place this sheet as the leftmost sheet
# in the workbook
ws2 = wb.create_sheet('Another', 0)

#label current sheet of workbook
ws.title = 'Mysheet'

# print(wb.sheetnames)

wb2 = load_workbook('regions.xlsx')

new_sheet = wb2.create_sheet('NewSheet')

#Place the current sheet in this variable
active_sheet = wb2.active 

cell = active_sheet['A1']
#shows the location of A1 in wb2's current sheet
# print(cell)
# #shows the value in A1 of the current sheet
# print(cell.value)
# x = 0
# for value in active_sheet['A':'C']:
#     print(value[x])
#     x += 1
#     print(value[x])
#Show the value in the 1st row and 3rd column
print(active_sheet[1][2].value)
to_be_changed_value = active_sheet[1][2]
active_sheet['C3'] = 'Looking pretty replaced'
# active_sheet['A1'] = 0
# wb2.save('newly_updated_regions.xlsx')